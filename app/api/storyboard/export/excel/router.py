from fastapi import APIRouter, Body, Response
from fastapi.responses import JSONResponse
from app.utils.mongo import mongo_instance
from app.utils.core import get_user_id
import logging
import requests
import io
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import json
import tempfile
from app.locales.translations import get_translation
import urllib.parse

excel_router = APIRouter()

@excel_router.post('/excel')
async def operator(projectId: str = Body(..., embed=True)):
    
    user_id = get_user_id()
    project_id = mongo_instance.generate_object_id(projectId)
  
    # 查询项目信息
    project = await mongo_instance.find_one(
        "projects",
        {
            "_id": project_id,
            "user_id": user_id,
            "status": "active"
        }
    )
        
    if not project:
        return JSONResponse(content={
            "code": 1,
            "data": None,
            "msg": get_translation("project_not_found"),
        })
        
    shots_data = []
        
    # 先构建角色信息字典
    roles_info = {}
    role_ids = project.get("role_ids", [])
    for role_id in role_ids:
        # 从roles集合中获取角色名称
        role_info = await mongo_instance.find_one(
            "roles",
            {"_id": role_id, "status": "active"}
        )
        
        if role_info:
            role_name = role_info.get("role_name", "")
            role_resource_id = role_info.get("selected_role_resource_id")
            role_resource_url = ""
            
            # 从role_resources集合中获取角色资源URL
            if role_resource_id:
                role_resource = await mongo_instance.find_one(
                    "role_resources",
                    {"_id": role_resource_id}
                )
                if role_resource:
                    role_resource_url = role_resource.get("resource_url", "")
            
            # 存储角色信息
            roles_info[str(role_id)] = {
                "role_name": role_name,
                "role_resource_url": role_resource_url,
            }
            
    shot_ids = project.get("shot_ids", [])
            
    # 遍历项目中的所有分镜ID
    for shot_id in shot_ids:
        shot = await mongo_instance.find_one(
            "shots",
            {
                "_id": shot_id,
                "status": "active"
            }
        )
        
        if not shot:
            continue
        
        shot_resource = await mongo_instance.find_one(
            "shot_resources",
            {"_id": shot.get("selected_shot_resource_id", "")}
        )
            
        # 构建分镜数据
        shot_item = {}
        
        # 设置基本信息
        shot_item["shot_id"] = str(shot["_id"])
        
        # 设置镜头资源信息
        shot_item["shot_resource"] = {
            "shot_resource_url": shot_resource.get("resource_url", ""),
            "is_HD": shot_resource.get("is_HD", False),
            "shot_resource_id": str(shot.get("selected_shot_resource_id", ""))
        }
        
        # 设置场景描述
        shot_item["scene_description"] = {
            "background": shot.get("background", ""),
            "characters": [{
                "role_id": str(role["role_id"]),
                "role_name": roles_info.get(str(role["role_id"]), {}).get("role_name", ""),
                "action_and_emotion": role.get("action_and_emotion", "")
            } for role in shot.get("selected_roles", [])]
        }
        
        # 设置台词信息
        shot_item["dialogues"] = [{
            "role_id": str(dialogue["role_id"]),
            "content": dialogue.get("content", ""),
            "role_name": get_translation("voiceover") if str(dialogue["role_id"]) == "voiceover" else roles_info.get(str(dialogue["role_id"]), {}).get("role_name", "")
        } for dialogue in shot.get("dialogues", [])]
        
        # 获取角色信息
        shot_item["main_characters"] = [{
            "role_id": str(role["role_id"]),
            "role_name": roles_info.get(str(role["role_id"]), {}).get("role_name", ""),
            "role_resource_url": roles_info.get(str(role["role_id"]), {}).get("role_resource_url", "")
        } for role in shot.get("selected_roles", [])]
        
        # 构建镜头参数对象
        shot_size = shot.get("shot_size", 1)
        camera_angle = shot.get("camera_angle", 0)
        shot_type = shot.get("shot_type", 0)
        
        # 获取镜头参数值
        size_values = get_translation("size_values")
        angle_values = get_translation("angle_values")
        type_values = get_translation("type_values")
        
        shot_item["shot_size"] = size_values[shot_size] if shot_size < len(size_values) else ""
        shot_item["camera_angle"] = angle_values[camera_angle] if camera_angle < len(angle_values) else ""
        shot_item["shot_type"] = type_values[shot_type] if shot_type < len(type_values) else ""
        shot_item["shot_time"] = shot.get("shot_time", 3)
        
        shots_data.append(shot_item)
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "分镜表"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 5  # 镜号
    ws.column_dimensions['B'].width = 20  # 画面
    ws.column_dimensions['C'].width = 25  # 画面描述
    ws.column_dimensions['D'].width = 10  # 台词
    ws.column_dimensions['E'].width = 15  # 主要人物
    ws.column_dimensions['F'].width = 8   # 景别
    ws.column_dimensions['G'].width = 10  # 机位角度
    ws.column_dimensions['H'].width = 10  # 摄影机运动
    ws.column_dimensions['I'].width = 8   # 时长
    
    # 设置表头
    headers = get_translation("excel_headers")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    # 设置边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 填充数据
    for row_num, shot in enumerate(shots_data, 2):
        # 镜号
        ws.cell(row=row_num, column=1).value = row_num - 1
        
        # 画面 - 下载图片并插入
        img_url = shot["shot_resource"]["shot_resource_url"]
        if img_url:
            try:
                response = requests.get(img_url)
                if response.status_code == 200:
                    img_data = io.BytesIO(response.content)
                    img = Image(img_data)
                    
                    # 调整图片大小
                    max_width = 150
                    max_height = 150
                    width_ratio = max_width / img.width
                    height_ratio = max_height / img.height
                    ratio = min(width_ratio, height_ratio)
                    
                    img.width = int(img.width * ratio)
                    img.height = int(img.height * ratio)
                    
                    # 计算单元格位置
                    cell = ws.cell(row=row_num, column=2)
                    ws.row_dimensions[row_num].height = max(img.height * 0.75, 100)
                    
                    # 添加图片到单元格
                    img.anchor = cell.coordinate
                    ws.add_image(img)
            except Exception as e:
                logging.error(f"Download or handle image Error: {str(e)}")
        
        # 画面描述
        scene_desc = shot["scene_description"]["background"]
        for character in shot["scene_description"]["characters"]:
            if character["action_and_emotion"]:
                scene_desc += f"\n{character['role_name']}: {character['action_and_emotion']}"
        ws.cell(row=row_num, column=3).value = scene_desc
        ws.cell(row=row_num, column=3).alignment = Alignment(wrapText=True, vertical='top')
        
        # 台词
        dialogues_text = ""
        for dialogue in shot["dialogues"]:
            if dialogue["content"]:
                dialogues_text += f"{dialogue['role_name']}: {dialogue['content']}\n"
        ws.cell(row=row_num, column=4).value = dialogues_text.strip()
        ws.cell(row=row_num, column=4).alignment = Alignment(wrapText=True, vertical='top')
        
        # 主要人物
        characters_text = ", ".join([char["role_name"] for char in shot["main_characters"]])
        ws.cell(row=row_num, column=5).value = characters_text
        ws.cell(row=row_num, column=5).alignment = Alignment(wrapText=True, vertical='center')
        
        # 景别
        ws.cell(row=row_num, column=6).value = shot["shot_size"]
        ws.cell(row=row_num, column=6).alignment = Alignment(horizontal='center', vertical='center')
        
        # 机位角度
        ws.cell(row=row_num, column=7).value = shot["camera_angle"]
        ws.cell(row=row_num, column=7).alignment = Alignment(horizontal='center', vertical='center')
        
        # 摄影机运动
        ws.cell(row=row_num, column=8).value = shot["shot_type"]
        ws.cell(row=row_num, column=8).alignment = Alignment(horizontal='center', vertical='center')
        
        # 时长
        second_text = get_translation("second")
        ws.cell(row=row_num, column=9).value = f"{shot['shot_time']}{second_text}"
        ws.cell(row=row_num, column=9).alignment = Alignment(horizontal='center', vertical='center')
        
        # 应用边框样式到所有单元格
        for col in range(1, 10):
            ws.cell(row=row_num, column=col).border = thin_border
    
    # 应用边框样式到表头
    for col in range(1, 10):
        ws.cell(row=1, column=col).border = thin_border
    
    # 保存到临时文件
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    
    # 读取文件内容
    with open(tmp_path, 'rb') as f:
        file_content = f.read()
    
    # 删除临时文件
    os.unlink(tmp_path)
    
    # 设置响应头，使浏览器下载文件
    project_name = project.get("project_name", "shots")
    filename = f"{project_name}.xlsx"
    encoded_filename = urllib.parse.quote(filename)
    
    headers = {
        'Content-Disposition': f'attachment; filename*=UTF-8\'\'{encoded_filename}',
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }
    
    return Response(content=file_content, headers=headers)